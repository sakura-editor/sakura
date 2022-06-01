﻿# -*- coding: utf-8 -*-
import sys
import re
import os
import csv
import platform
import appveyor_env

# 解析結果を格納するハッシュのキー
logHashKeys = [
	'errType',
	'code',
	'source',
	'dest',
	'path',
	'lineNumber',
	'message',
	'relpath',
	'blobURL',
]

csvKeys = [
	'errType',
	'code',
	'source',
	'dest',
	'path',
	'lineNumber',
	'message',
	'blobURL',
]

excelKeys = [
	'errType',
	'code',
	'source',
	'dest',
	'relpath',
	'message',
	'path',
	'lineNumber',
]

colors_YELLOW = "FFFF00"
colors_BLUE   = "0000FF"

# infile: msbuild のビルドログ・ファイル名
# 戻り値: ログの解析結果が入ったハッシュの配列
def parse_buildlog(infile):
	# ファイル名に対する正規表現: 例 c:\hogehoge\hoge.c
	regLilePath  = r'(?P<filePath>[a-zA-Z]:([^(]+))'

	# 行番号に対する正規表現: 例 (100)
	regLineNumer = r'\((?P<lineNumber>\d+)\)'

	# エラーコードに対する正規表現: 例 warning C4267
	regError     = r'\s*(?P<errType>\w+)\s+(?P<code>\w+)\s*'

	# エラーメッセージ: 例 'argument': conversion from 'size_t' to 'int', possible loss of data [C:\projects\sakura\sakura\sakura.vcxproj]
	regMessage   = r'(?P<message>.+)$'

	# 解析対象の正規表現
	# 例
	#   c:\projects\sakura\sakura_core\basis\cmystring.h(39): warning C4267: 'argument': conversion from 'size_t' to 'int', possible loss of data [C:\projects\sakura\sakura\sakura.vcxproj]
	regEx        = regLilePath + regLineNumer + r':' + regError + r':' + regMessage

	# 型変換に対する警告メッセージ部分に対する正規表現: 例 'size_t' to 'int'
	regFromTo    = r"from '(?P<source>[^']+)' to '(?P<dest>[^']+)'"

	data = []
	with open(infile, "r") as fin:
		# msbuild-xxx-xxx.log のログに警告が重複して出現することに対する Workaround 用
		duplicateCheck = {}

		appveyor = appveyor_env.AppveyorEnv()

		print ("open " + infile)
		for line in fin:
			text = line.replace('\n','')
			text = text.replace('\r','')

			match = re.search(regEx, text)
			if match:
				path       = match.group('filePath')
				lineNumber = match.group('lineNumber')
				errType    = match.group('errType')
				code       = match.group('code')
				message    = match.group('message')

				entry = {}
				entry['errType'] = errType
				entry['code'] = code

				match2 = re.search(regFromTo, text)
				if match2:
					source  = match2.group('source')
					dest    = match2.group('dest')
					entry['source'] = source
					entry['dest']   = dest
				else:
					entry['source'] = r''
					entry['dest']   = r''

				entry['path']       = path
				entry['lineNumber'] = lineNumber
				entry['message']    = message
				
				(blobURL, relpath) = appveyor.getBlobURLWithLine(path, lineNumber)
				entry['relpath'] = relpath
				entry['blobURL'] = blobURL

				temp = []
				for key in logHashKeys:
					temp.append(entry[key])
				entry['key'] = ' '.join(temp)

				# msbuild-xxx-xxx.log のログに同じ警告が重複して出現することに対する Workaround
				# entry['key'] で警告を一意に識別できるので同じ値の場合に CSV に出力しない
				logKey = entry['key']
				duplicateCheck[logKey] = duplicateCheck.get(logKey, 0) + 1
				if duplicateCheck[logKey] == 1:
					data.append(entry)

	# ソート対象のハッシュ配列で 'key' というキーを元にソートする。
	from operator import itemgetter
	data = sorted(data, key=itemgetter('key'))

	return data

# outfile: 出力ファイル名
# data   : ログの解析結果が入ったハッシュの配列 ( parse_buildlog() の戻り値 )
def writeToCSV(outfile, data):
	# 解析結果を CSV ファイルに出力する
	with open(outfile, "w") as fout:
		writer = csv.writer(fout, lineterminator='\n')
		writer.writerow(csvKeys)

		for entry in data:
			temp = []
			for key in csvKeys:
				temp.append(entry[key])
			writer.writerow(temp)

		print ("wrote " + outfile)
		
# outfile: 出力ファイル名
# data   : ログの解析結果が入ったハッシュの配列 ( parse_buildlog() の戻り値 )
def writeToXLSX(outfile, data):
	# CELL に設定する値を変換する関数を返す
	def getEntryConverter():
		def converterPython3(value):
			return value
	
		def converterPython2(value):
			return value.decode('shiftjis').encode('utf_8')

		if sys.version_info.major >= 3:
			return converterPython3

		return converterPython2

	try:
		import openpyxl
		from openpyxl.styles import colors
		from openpyxl.styles import Font, Color
		from openpyxl.styles.fills import PatternFill

		wb = openpyxl.Workbook()
		ws = wb.active

		# 列幅に必要なサイズを保持する配列
		maxWidths = []
		
		# ヘッダー部分を設定する
		y = 0
		for x, item in enumerate(excelKeys):
			cell = ws.cell(row=y+1, column=x+1)
			cell.value = item
			cell.fill  = PatternFill(patternType='solid', start_color=colors_YELLOW, end_color=colors_YELLOW)
			maxWidths.append(len(cell.value) + 1)
		y = y + 1

		# 各エントリーを設定するときのコンバーターを取得する (python 2/3 の違いを吸収するためのもの)
		converter = getEntryConverter()
		
		# エラーごとにまとめて表示するための情報
		errorSummary = {}

		# ログの解析結果を設定する
		for entry in data:
			# ファイルパスを取り除いた関連エラーメッセージをまとめる
			message = converter(entry['message'])
			message = re.sub(r'((\S*)\)\s*)?\[(.+?)\]', r'', message)
			
			# エラータイプ
			errType = entry['errType']
			
			# エラーコード
			code = entry['code']
			
			# サマリーを管理するハッシュ用のキー
			errorKey = ' '.join([errType, code, message])

			if errorKey not in errorSummary:
				errorSummary[errorKey] = {}
				errorSummary[errorKey]["description"] = message
				errorSummary[errorKey]["entries"]     = []
				errorSummary[errorKey]["errType"]     = errType
				errorSummary[errorKey]["code"]        = code
			errorSummary[errorKey]["entries"].append(entry)

			for x, key in enumerate(excelKeys):
				cell = ws.cell(row=y+1, column=x+1)
				if key == "relpath":
					val            = entry['relpath'] + " line: " + entry['lineNumber']
					cell.hyperlink = entry['blobURL']
					cell.font      = Font(u='single', color=colors_BLUE)
				else:
					entryKey = entry[key]
					val  = converter(entry[key])

				# 列幅を設定するために必要なサイズを計算する
				width = len(val) + 1
				if maxWidths[x] < width:
					maxWidths[x] = width
					
				# セルに値を設定する
				if val.isdigit():
					cell.value = int(val)
				else:
					cell.value = val

			# 行番号を更新する
			y = y + 1

		# 列幅を設定する
		for x, item in enumerate(excelKeys):
			ws.column_dimensions[openpyxl.utils.get_column_letter(x+1)].width = maxWidths[x]
		
		# Excel の列にフィルタを設定する
		start = openpyxl.utils.get_column_letter(1)
		end   = openpyxl.utils.get_column_letter(len(excelKeys))
		ws.auto_filter.ref = start + ":" + end
		
		# ウィンドウ枠を固定
		ws.freeze_panes = 'F2'

		#############################################################################
		#	エラーのサマリーシート用のコード
		#############################################################################
		worksheetIndex = 0
		errorKeys = sorted(errorSummary.keys())
		for errorKey in errorKeys:
			worksheetIndex = worksheetIndex + 1

			# 列幅に必要なサイズを保持する配列
			maxWidths = []

			message = errorSummary[errorKey]["description"]
			entries = errorSummary[errorKey]["entries"]
			errType = errorSummary[errorKey]["errType"]
			code    = errorSummary[errorKey]["code"]

			wsError = wb.create_sheet()
			wsError.title = str(worksheetIndex) + "(" + code + ")"
			
			x = 0
			y = 0
			
			# エラーメッセージをセットする
			if entries:
				entry = entries[0]
				cell = wsError.cell(row=y+1, column=x+1)
				cell.value = message
			y = y + 1
			
			# 空行を入れる
			y = y + 1
			
			outputKeys = [
				'path',
				'lineNumber',
				'blobURL',
			]
		
			# ヘッダー部分を設定
			for x, key in enumerate(outputKeys):
				cell = wsError.cell(row=y+1, column=x+1)
				cell.value = key
				cell.fill  = PatternFill(patternType='solid', start_color=colors_YELLOW, end_color=colors_YELLOW)
				maxWidths.append(len(cell.value) + 1)
			y = y + 1

			# 各エラー箇所を設定する
			for entry in entries:
				for x, key in enumerate(outputKeys):
					cell = wsError.cell(row=y+1, column=x+1)
					if key == "blobURL":
						cell.hyperlink = entry[key]
						cell.font      = Font(u='single', color=colors_BLUE)
					val = converter(entry[key])
					if val.isdigit():
						cell.value = int(val)
					else:
						cell.value = val

					# 列幅を設定するために必要なサイズを計算する
					width = len(val) + 1
					if maxWidths[x] < width:
						maxWidths[x] = width

				y = y + 1

			# 列幅を設定する
			for x, item in enumerate(maxWidths):
				wsError.column_dimensions[openpyxl.utils.get_column_letter(x+1)].width = item

		wb.save(outfile)
		print ("wrote " + outfile)
	except ImportError:
		print ("please run '<python root>\\Scripts\\pip install openpyxl --user'")

# main 関数
def main():
	if len(sys.argv) < 2:
		print ("usage: " + sys.argv[0] + " <logfile name>")
		sys.exit(1)

	infile = sys.argv[1]
	outcsvfile = infile + '.csv'
	outxlsxfile = infile + '.xlsx'

	data = parse_buildlog(infile)
	writeToCSV(outcsvfile, data)
	writeToXLSX(outxlsxfile, data)

if __name__ == '__main__':
	main()
